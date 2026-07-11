"""
Figure 19: BIPV Color Variation Life Cycle Cost (LCC) Analysis
==============================================================
This script evaluates the economic performance of Case 3 (Kinetic BIPV) color variations:
- Case 1: Base (No BIPV)
- Case 3 — White (10.51% efficiency)
- Case 3 — Light Gray Beige (14.71% efficiency)
- Case 3 — Terracotta (16.39% efficiency)

It calculates and plots:
1) Nominal Cumulative Cost Comparison (Undiscounted)
2) Life Cycle Cost (LCC) Comparison (Discounted Present Value)
"""

import os
import re
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import dartwork_mpl as dm
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# Save SVGs as vector paths for Illustrator/Figma editability
plt.rcParams['svg.fonttype'] = 'none'

# ── Paths ─────────────────────────────────────────────────────────────
script_dir  = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.normpath(os.path.join(script_dir, "..", ".."))
script_name = os.path.splitext(os.path.basename(__file__))[0]
figure_dir = os.path.normpath(os.path.join(script_dir, "..", "plot", script_name))
excel_path  = os.path.join(project_root, "economy_analysis.xlsx")
data_dir     = os.path.join(project_root, "case_analysis", "bipv_variation")
os.makedirs(figure_dir, exist_ok=True)

# ── Load PV Generation from CSVs ──────────────────────────────────────
def get_annual_pv_mwh(csv_filename):
    csv_path = os.path.join(data_dir, csv_filename)
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"PV data file not found: {csv_path}")
    df = pd.read_csv(csv_path, usecols=['Whole Building:Facility Total Produced Electricity Energy [J](TimeStep)'])
    joules = df['Whole Building:Facility Total Produced Electricity Energy [J](TimeStep)'].sum()
    return joules / 3.6e9  # Joules to MWh

print("Loading annual PV generation data...")
PV0_MWH = {
    1: 0.0,
    "White": get_annual_pv_mwh("case3_white.csv"),
    "Beige": get_annual_pv_mwh("case3_light_gray_beige.csv"),
    "Terracotta": get_annual_pv_mwh("case3_terracotta.csv")
}
for k, v in PV0_MWH.items():
    print(f"  {k}: {v:.3f} MWh")

# ── Excel Parameter Reading ───────────────────────────────────────────
wb  = load_workbook(excel_path, data_only=False)
ws3 = wb["가정"]          # Case 3 (Kinetic)
ws1 = wb["가정_Case1"]    # Case 1 (Base)

def eval_cell(ws, row, col=2):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str) and v.startswith("="):
        expr = v[1:]
        m = re.fullmatch(r'([A-Z]+)(\d+)', expr)
        if m:
            return eval_cell(ws, int(m.group(2)), column_index_from_string(m.group(1)))
        try:
            return float(eval(expr))
        except Exception:
            return 0.0
    return 0.0 if v is None else float(v)

# Common Economic Parameters (using ws3 as baseline)
r_deg    = eval_cell(ws3, 7)    # PV degradation rate = 0.005
n_years  = int(eval_cell(ws3, 8))  # Project lifetime = 25
r_elec   = eval_cell(ws3, 10)   # Tariff escalation = 0.05
r_inf    = eval_cell(ws3, 11)   # Inflation rate = 0.02
r_disc   = eval_cell(ws3, 12)   # Discount rate = 0.055

# Weighted average electricity tariff calculation
b42 = eval_cell(ws3, 42)   # 기후환경요금
b43 = eval_cell(ws3, 43)   # 연료비조정요금
b44 = eval_cell(ws3, 44)   # 부가가치세
b45 = eval_cell(ws3, 45)   # 전력기금
d34 = eval_cell(ws3, 34, col=4)   # 여름 비중
d35 = eval_cell(ws3, 35, col=4)   # 봄가을 비중
d36 = eval_cell(ws3, 36, col=4)   # 겨울 비중
factor = 1 + b44 + b45
tariff_0 = ((132.4 + b42 + b43) * factor * d34
          + (91.9  + b42 + b43) * factor * d35
          + (119.0 + b42 + b43) * factor * d36)

# CAPEX and O&M parameters
def calc_capex(ws):
    return sum(eval_cell(ws, r) for r in range(15, 23))

capex_c3 = calc_capex(ws3)
om_rate_c3 = eval_cell(ws3, 26)
rep_year_c3 = int(eval_cell(ws3, 28))
rep_cost_c3 = eval_cell(ws3, 29)
drive_3_kwh = eval_cell(ws3, 6)   # Case 3 actuator drive energy (kWh/year)

CASES = [1, "White", "Beige", "Terracotta"]

# Map economic parameters to cases
capex_map = {1: 0.0, "White": capex_c3, "Beige": capex_c3, "Terracotta": capex_c3}
om_yr1_map = {1: 0.0, "White": capex_c3 * om_rate_c3, "Beige": capex_c3 * om_rate_c3, "Terracotta": capex_c3 * om_rate_c3}
rep_year_map = {1: 99, "White": rep_year_c3, "Beige": rep_year_c3, "Terracotta": rep_year_c3}
rep_cost_map = {1: 0.0, "White": rep_cost_c3, "Beige": rep_cost_c3, "Terracotta": rep_cost_c3}

# HVAC annual electric consumption (MWh)
# Only PV generation varies; HVAC load remains the same as Kinetic BIPV (Case 3) for all variations.
HVAC_MWH = {
    1: 48.994,
    "White": 48.722,
    "Beige": 48.722,
    "Terracotta": 48.722
}

# ── Annual Cost Calculations ──────────────────────────────────────────
years = np.arange(0, n_years + 1)
annual_cost = {k: np.zeros(n_years + 1) for k in CASES}

for k in CASES:
    annual_cost[k][0] = capex_map[k]

for t in range(1, n_years + 1):
    tariff_t = tariff_0 * (1 + r_elec) ** (t - 1)
    for k in CASES:
        pv_t      = PV0_MWH[k] * (1 - r_deg) ** (t - 1)
        net_pv    = pv_t - (drive_3_kwh / 1000.0 if k != 1 else 0.0)
        net_kWh   = (HVAC_MWH[k] - net_pv) * 1000.0
        elec_cost = net_kWh * tariff_t
        om_t      = om_yr1_map[k] * (1 + r_inf) ** (t - 1)
        rep_t     = rep_cost_map[k] * (1 + r_inf) ** t if (t % rep_year_map[k] == 0) else 0.0
        annual_cost[k][t] = elec_cost + om_t + rep_t

# ── Cumulative Nominal Cost (Undiscounted) ───────────────────────────
cum_nominal = {k: np.zeros(n_years + 1) for k in CASES}
for k in CASES:
    cum_nominal[k][0] = annual_cost[k][0]
    for t in range(1, n_years + 1):
        cum_nominal[k][t] = cum_nominal[k][t - 1] + annual_cost[k][t]

# ── Life Cycle Cost (LCC, Discounted Present Value) ──────────────────
cum_lcc = {k: np.zeros(n_years + 1) for k in CASES}
for k in CASES:
    cum_lcc[k][0] = annual_cost[k][0]   # CAPEX is undiscounted at Year 0
    for t in range(1, n_years + 1):
        cum_lcc[k][t] = cum_lcc[k][t - 1] + annual_cost[k][t] / (1 + r_disc) ** t

# Convert to Million KRW (M KRW)
for k in CASES:
    cum_nominal[k] /= 1e6
    cum_lcc[k]     /= 1e6

print("\n25-Year Nominal Cumulative Cost:")
for k in CASES:
    print(f"  {k}: {cum_nominal[k][25]:.1f} M KRW")
print("\n25-Year Life Cycle Cost (LCC):")
for k in CASES:
    print(f"  {k}: {cum_lcc[k][25]:.1f} M KRW")

# ── Plot Styling ──────────────────────────────────────────────────────
dm.style.use("presentation")
plt.rcParams.update({"xtick.labelsize": 12, "ytick.labelsize": 12})

colors  = {1: "oc.gray7", "White": "#64748b", "Beige": "#a8a29e", "Terracotta": "#c2593f"}
markers = {1: "o",        "White": "D",       "Beige": "s",       "Terracotta": "^"}
labels  = {
    1:            "Case 1 (Base, No BIPV)",
    "White":      "Case 3 — White (10.51%)",
    "Beige":      "Case 3 — Light Gray Beige (14.71%)",
    "Terracotta": "Case 3 — Terracotta (16.39%)"
}
lws = {1: 1.5, "White": 1.5, "Beige": 1.5, "Terracotta": 1.5}

def compute_label_positions(cum, min_gap=8.0):
    sorted_cases = sorted(CASES, key=lambda k: cum[k][25])
    pos = {k: float(cum[k][25]) for k in sorted_cases}

    for _ in range(300):
        changed = False
        for i in range(1, len(sorted_cases)):
            k, k_prev = sorted_cases[i], sorted_cases[i - 1]
            gap = pos[k] - pos[k_prev]
            if gap < min_gap:
                push = (min_gap - gap) / 2
                pos[k_prev] -= push
                pos[k]      += push
                changed = True
        if not changed:
            break
    return pos

def draw_cost_plot(ax, cum, ylabel, title):
    for k in CASES:
        ax.plot(years, cum[k], color=colors[k], lw=dm.lw(lws[k]),
                marker=markers[k], markersize=4, label=labels[k])

    # Automatic label positioning at year 25
    label_y = compute_label_positions(cum, min_gap=max(8.0, (max(cum[k].max() for k in CASES) * 0.05)))
    y_top   = max(max(cum[k].max() for k in CASES), max(label_y.values())) * 1.06
    ax.set_ylim(0, y_top)

    for k in CASES:
        ax.annotate(f"{cum[k][25]:.1f} M",
                    xy=(25, cum[k][25]), xytext=(25.5, label_y[k]),
                    fontsize=9.5, color=colors[k], fontweight="bold", va="center",
                    arrowprops=dict(arrowstyle="-", color=colors[k], lw=0.9))

    # Year 12 & 24 replacement event lines
    ax.axvline(12, color="oc.gray5", linestyle=":", lw=dm.lw(0.8))
    ax.axvline(24, color="oc.gray5", linestyle=":", lw=dm.lw(0.8))
    ax.text(12.3, y_top * 0.03,
            "Year 12\n(Actuator\nReplacement)",
            color="oc.gray6", fontsize=9, va="bottom")
    ax.text(24.3, y_top * 0.03,
            "Year 24\n(Actuator\nReplacement)",
            color="oc.gray6", fontsize=9, va="bottom")

    ax.set_xlabel("Project Timeline [Years]")
    ax.set_ylabel(ylabel)
    ax.set_title(title, fontweight="bold")
    ax.set_xlim(0, 28)
    ax.set_xticks(range(0, 26, 5))
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}"))
    ax.legend(loc="upper left")

# ── Plot 1: Nominal Cumulative Cost ──────────────────────────────────
fig, ax = plt.subplots(figsize=(23 / 2.54, 14 / 2.54))
draw_cost_plot(
    ax, cum_nominal,
    ylabel="Nominal Cumulative Cost [Million KRW]",
    title="Nominal Cumulative Cost Comparison by BIPV Color (Case 3)\n"
          "(CAPEX + HVAC Electricity + O&M + Replacement, Undiscounted)",
)
dm.simple_layout(fig)
fig.savefig(os.path.join(figure_dir, "fig19_cumulative_nominal_color.png"), dpi=300, transparent=True)
fig.savefig(os.path.join(figure_dir, "fig19_cumulative_nominal_color.svg"), transparent=True)
plt.close(fig)

# ── Plot 2: LCC (Discounted Present Value) ───────────────────────────
fig, ax = plt.subplots(figsize=(23 / 2.54, 14 / 2.54))
draw_cost_plot(
    ax, cum_lcc,
    ylabel="Life Cycle Cost — Present Value [Million KRW]",
    title=f"Life Cycle Cost (LCC) Comparison by BIPV Color (Case 3)\n"
          f"(Discount Rate {r_disc*100:.1f}%,  Tariff Escalation {r_elec*100:.0f}%/yr,  PV Degradation {r_deg*100:.1f}%/yr)",
)
dm.simple_layout(fig)
fig.savefig(os.path.join(figure_dir, "fig19_cumulative_lcc_color.png"), dpi=300, transparent=True)
fig.savefig(os.path.join(figure_dir, "fig19_cumulative_lcc_color.svg"), transparent=True)
plt.close(fig)

print(f"\nSaved plots to: {figure_dir}")
