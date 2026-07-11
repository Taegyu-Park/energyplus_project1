"""
Figure 20: BIPV Color Variation DCF-based Payback Comparison
=============================================================
This script calculates and compares the Discounted Cash Flow (DCF)-based
investment recovery (payback period) for Case 3 (Kinetic BIPV) color variations:
- Case 3 — White (10.51% efficiency)
- Case 3 — Light Gray Beige (14.71% efficiency)
- Case 3 — Terracotta (16.39% efficiency)

All parameters (CAPEX, O&M rate, replacement year, replacement cost)
are identical, and only PV generation varies by color.
"""

import os
import re
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import dartwork_mpl as dm
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# Save SVGs as vector paths for Illustrator/Figma editability
plt.rcParams['svg.fonttype'] = 'none'

# ── Paths ─────────────────────────────────────────────────────────────
script_dir   = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.normpath(os.path.join(script_dir, "..", ".."))
script_name  = os.path.splitext(os.path.basename(__file__))[0]
figure_dir   = os.path.normpath(os.path.join(script_dir, "..", "plot", script_name))
excel_path   = os.path.join(project_root, "economy_analysis.xlsx")
data_dir     = os.path.join(project_root, "case_analysis", "bipv_variation")
os.makedirs(figure_dir, exist_ok=True)

# ── Load PV Generation from CSVs (kWh/yr) ──────────────────────────────
def get_annual_pv_kwh(csv_filename):
    csv_path = os.path.join(data_dir, csv_filename)
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"PV data file not found: {csv_path}")
    df = pd.read_csv(csv_path, usecols=['Whole Building:Facility Total Produced Electricity Energy [J](TimeStep)'])
    joules = df['Whole Building:Facility Total Produced Electricity Energy [J](TimeStep)'].sum()
    return joules / 3.6e6  # Joules to kWh

print("Loading annual PV generation data (kWh)...")
PV_kwh = {
    "White": get_annual_pv_kwh("case3_white.csv"),
    "Beige": get_annual_pv_kwh("case3_light_gray_beige.csv"),
    "Terracotta": get_annual_pv_kwh("case3_terracotta.csv")
}
for k, v in PV_kwh.items():
    print(f"  {k}: {v:.1f} kWh")

# ── Excel Parameter Reading ───────────────────────────────────────────
wb  = load_workbook(excel_path, data_only=False)
ws3 = wb["가정"]

def ec(ws, row, col=2):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str) and v.startswith("="):
        expr = v[1:]
        m = re.fullmatch(r'([A-Z]+)(\d+)', expr)
        if m:
            return ec(ws, int(m.group(2)), column_index_from_string(m.group(1)))
        try: return float(eval(expr))
        except: return 0.0
    return 0.0 if v is None else float(v)

# Common Economic Parameters
b42, b43, b44, b45 = ec(ws3,42), ec(ws3,43), ec(ws3,44), ec(ws3,45)
d34, d35, d36       = ec(ws3,34,4), ec(ws3,35,4), ec(ws3,36,4)
f = 1 + b44 + b45
tariff_0 = ((132.4+b42+b43)*f*d34 + (91.9+b42+b43)*f*d35 + (119+b42+b43)*f*d36)

r_elec  = ec(ws3, 10)   # 0.05
r_inf   = ec(ws3, 11)   # 0.02
r_deg   = ec(ws3, 7)    # 0.005
r_disc  = ec(ws3, 12)   # 0.055
n_years = int(ec(ws3, 8))  # 25

# CAPEX & Operating costs for Case 3
capex3  = sum(ec(ws3, r) for r in range(15, 23))
om_r3   = ec(ws3, 26)
rep_yr3 = int(ec(ws3, 28))
rep_c3  = ec(ws3, 29)
drive3  = ec(ws3, 6)   # kWh/yr

# ── DCF Calculation ───────────────────────────────────────────────────
def calc_dcf(capex, pv0_kwh, drive_kwh, om_rate, rep_year, rep_cost_now):
    cum  = np.zeros(n_years + 1)
    cum[0] = -capex
    om0  = capex * om_rate
    for t in range(1, n_years + 1):
        pv_t    = pv0_kwh * (1 - r_deg) ** (t - 1)
        net_pv  = pv_t - drive_kwh
        savings = net_pv  * tariff_0 * (1 + r_elec) ** (t - 1)
        om_t    = om0     * (1 + r_inf)  ** (t - 1)
        rep_t   = rep_cost_now * (1 + r_inf) ** t if (t % rep_year == 0) else 0.0
        cum[t]  = cum[t - 1] + (savings - om_t - rep_t) / (1 + r_disc) ** t
    return cum

CASES = ["White", "Beige", "Terracotta"]
dcf = {k: calc_dcf(capex3, PV_kwh[k], drive3, om_r3, rep_yr3, rep_c3) for k in CASES}
dcf_M = {k: dcf[k] / 1e6 for k in CASES}
years = np.arange(n_years + 1)

# Payback calculations (linear interpolation)
def payback_yr(arr):
    for i in range(1, len(arr)):
        if arr[i - 1] < 0 <= arr[i]:
            return (i - 1) + abs(arr[i - 1]) / (arr[i] - arr[i - 1])
    return None

pb = {k: payback_yr(dcf_M[k]) for k in CASES}

for k in CASES:
    pb_val = pb[k]
    pb_str = f"{pb_val:.1f} yr" if pb_val is not None else "> 25 yr"
    print(f"[{k}] Payback: {pb_str:>8s} | 25yr NPV: {dcf_M[k][-1]:+.2f} M KRW")

# ── Plotting ─────────────────────────────────────────────────────────
dm.style.use("presentation")
plt.rcParams.update({"xtick.labelsize": 12, "ytick.labelsize": 12})

fig, ax = plt.subplots(figsize=(26 / 2.54, 14 / 2.54))

colors = {"White": "#64748b", "Beige": "#a8a29e", "Terracotta": "#c2593f"}
labels = {
    "White": "Case 3 — White (10.51%)",
    "Beige": "Case 3 — Light Gray Beige (14.71%)",
    "Terracotta": "Case 3 — Terracotta (16.39%)"
}

# Y axis limits and styling
all_vals = np.concatenate(list(dcf_M.values()))
y_lo = all_vals.min() * 1.15
y_hi = all_vals.max() * 1.25
ax.set_ylim(y_lo, y_hi)

# Shading negative/positive regions
ax.fill_between(years, y_lo, 0, color="oc.red2",  alpha=0.08, linewidth=0, zorder=0)
ax.fill_between(years, 0, y_hi, color="oc.teal1", alpha=0.10, linewidth=0, zorder=0)
ax.axhline(0, color="oc.gray5", linestyle="--", lw=dm.lw(0.8), zorder=1)

# Plot curves
for k in CASES:
    ax.plot(years, dcf_M[k], color=colors[k], lw=dm.lw(1.5), label=labels[k], zorder=3)

# Payback markers and lines
# To avoid text overlapping, we will place text labels at different heights or offset positions
pb_y_positions = {"White": y_hi * 0.88, "Beige": y_hi * 0.70, "Terracotta": y_hi * 0.52}
for k in CASES:
    p_yr = pb[k]
    if p_yr is not None:
        ax.axvline(p_yr, color=colors[k], linestyle=":", lw=dm.lw(1.2), zorder=2)
        ax.plot(p_yr, 0, marker="o", color=colors[k], markersize=7, zorder=5)
        ax.text(p_yr - 0.3, pb_y_positions[k],
                f"{k} Payback\n{p_yr:.1f} yr",
                color=colors[k], fontsize=10, fontweight="bold", ha="right", va="top")

# Actuator replacement indicator on Year 12 & 24 (only label one curve to avoid clutter)
for k in CASES:
    ax.plot(12, dcf_M[k][12], marker="x", color=colors[k], markersize=9, mew=2, zorder=5)
    ax.plot(24, dcf_M[k][24], marker="x", color=colors[k], markersize=9, mew=2, zorder=5)

# Add arrow pointing to Terracotta Year 12 replacement
ax.annotate("Actuator\nReplacement\n(Year 12 & 24)",
            xy=(12, dcf_M["Terracotta"][12]), xytext=(9.0, -10),
            fontsize=9.5, color="oc.orange9", ha="right", va="center",
            arrowprops=dict(arrowstyle="->", color="oc.orange9", lw=1.1))

# Right end NPV labels (with automatic spacing to avoid overlap)
npv_vals = {k: dcf_M[k][-1] for k in CASES}
def compute_right_label_positions(npvs, min_gap=2.2):
    sorted_k = sorted(npvs.keys(), key=lambda k: npvs[k])
    pos = {k: float(npvs[k]) for k in sorted_k}
    for _ in range(300):
        changed = False
        for i in range(1, len(sorted_k)):
            k, k_prev = sorted_k[i], sorted_k[i - 1]
            gap = pos[k] - pos[k_prev]
            if gap < min_gap:
                push = (min_gap - gap) / 2
                pos[k_prev] -= push
                pos[k]      += push
                changed = True
        if not changed:
            break
    return pos

label_y = compute_right_label_positions(npv_vals)
for k in CASES:
    pb_val = pb[k]
    suffix = f"  ({pb_val:.1f} yr)" if pb_val is not None else "  (> 25 yr)"
    ax.text(25.3, label_y[k],
            f"{dcf_M[k][-1]:+.1f} M  ({k}){suffix}",
            color=colors[k], fontsize=10, fontweight="bold", va="center")

# Title and axis formatting
ax.set_xlabel("Project Timeline [Years]")
ax.set_ylabel("Cumulative NPV [Million KRW]")
ax.set_title(
    "DCF-based Investment Recovery Comparison by BIPV Color (Case 3)\n"
    f"(Discount Rate {r_disc*100:.1f}%,  Tariff Escalation {r_elec*100:.0f}%/yr,  PV Degradation {r_deg*100:.1f}%/yr)",
    fontweight="bold",
)
ax.set_xlim(0, 31)
ax.set_xticks(range(0, 26, 5))
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}"))
ax.legend(loc="lower right", fontsize=9.5, ncol=1)

# Save figure
output_png = os.path.join(figure_dir, "dcf_payback_comparison_color.png")
output_svg = os.path.join(figure_dir, "dcf_payback_comparison_color.svg")

dm.simple_layout(fig)
fig.savefig(output_png, dpi=300, transparent=True)
fig.savefig(output_svg, transparent=True)
plt.close(fig)

print(f"\nSaved payback plot to: {figure_dir}")
