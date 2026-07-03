"""
Cumulative Total Cost of Ownership (TCO) Comparison
- Case 1: Base (BIPV 없음, 투자 없음)
- Case 2: Static BIPV 90°
- Case 3: Kinetic BIPV

TCO = CAPEX + Σ_t (HVAC 전기요금 + O&M + 교체비)
  - HVAC 전기 = 냉난방 열부하 / COP  (냉방 COP=3, 난방 COP=2.5)
  - PV 발전량으로 전기요금 상계 (순발전 × 전기요금)
  - 전기요금: 가중평균 tariff, 연 5% 상승
  - PV 발전량: 연 0.5% 열화
"""

import os
import re
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import dartwork_mpl as dm
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ── 경로 ─────────────────────────────────────────────────────────────
script_dir  = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.normpath(os.path.join(script_dir, "..", ".."))
script_name = os.path.splitext(os.path.basename(__file__))[0]
figure_dir = os.path.normpath(os.path.join(script_dir, "..", "plot", script_name))
excel_path  = os.path.join(project_root, "economy_analysis.xlsx")
os.makedirs(figure_dir, exist_ok=True)

# ── Excel 읽기 ────────────────────────────────────────────────────────
wb  = load_workbook(excel_path, data_only=False)
ws3 = wb["가정"]          # Case 3 (Kinetic)
ws2 = wb["가정_Case2"]    # Case 2 (Static 90°)
ws1 = wb["가정_Case1"]    # Case 1 (Base)

def eval_cell(ws, row, col=2):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str) and v.startswith("="):
        expr = v[1:]
        m = re.fullmatch(r'([A-Z]+)(\d+)', expr)
        if m:
            return eval_cell(ws, int(m.group(2)), column_index_from_string(m.group(1)))
        try:
            return float(eval(expr))
        except Exception:
            return 0.0
    return 0.0 if v is None else float(v)

# 공통 경제 가정 (가정 = Case 3 시트 기준)
r_deg    = eval_cell(ws3, 7)    # B7:  PV 열화율   = 0.005
n_years  = int(eval_cell(ws3, 8))  # B8:  사업기간    = 25
r_elec   = eval_cell(ws3, 10)   # B10: 전기요금 상승률 = 0.05
r_inf    = eval_cell(ws3, 11)   # B11: 물가상승률  = 0.02

# 전기요금 가중평균 tariff 계산
b42 = eval_cell(ws3, 42)   # 기후환경요금
b43 = eval_cell(ws3, 43)   # 연료비조정요금
b44 = eval_cell(ws3, 44)   # 부가가치세
b45 = eval_cell(ws3, 45)   # 전력기금
d34 = eval_cell(ws3, 34, col=4)   # 여름 비중
d35 = eval_cell(ws3, 35, col=4)   # 봄가을 비중
d36 = eval_cell(ws3, 36, col=4)   # 겨울 비중
factor = 1 + b44 + b45
tariff_0 = ((132.4 + b42 + b43) * factor * d34
          + (91.9  + b42 + b43) * factor * d35
          + (119.0 + b42 + b43) * factor * d36)   # ≈ 138.7 원/kWh

# CAPEX
def calc_capex(ws):
    return sum(eval_cell(ws, r) for r in range(15, 23))

capex   = {1: 0.0, 2: calc_capex(ws2), 3: calc_capex(ws3)}
om_rate = {1: 0.0, 2: eval_cell(ws2, 26), 3: eval_cell(ws3, 26)}
om_yr1  = {k: capex[k] * om_rate[k] for k in [1, 2, 3]}

rep_year = {1: 99, 2: int(eval_cell(ws2, 28)), 3: int(eval_cell(ws3, 28))}
rep_cost = {1: 0.0, 2: eval_cell(ws2, 29), 3: eval_cell(ws3, 29)}

drive_3_kwh = eval_cell(ws3, 6)   # B6: Case 3 구동 소비전력 (kWh/year)

print(f"tariff_0      = {tariff_0:.2f} 원/kWh")
print(f"Case 2 CAPEX  = {capex[2]/1e4:.0f} 만원,  O&M = {om_rate[2]*100:.1f}%/년")
print(f"Case 3 CAPEX  = {capex[3]/1e4:.0f} 만원,  O&M = {om_rate[3]*100:.1f}%/년")
print(f"Case 3 교체   = 12년차 {rep_cost[3]/1e4:.0f} 만원 (현재가)")

# ── 시뮬레이션 기반 연간 HVAC 전력 소비 ──────────────────────────────
# EnergyPlus 냉난방 열부하 ÷ COP → 전기 소비  (MWh/year)
# 냉방 COP = 3.0,  난방 COP = 2.5
HVAC_MWH = {
    1:    48.994,
    "2_70": 51.460, "2_80": 50.994, "2_90": 50.110,
    3:    48.722,
}

# 1년차 PV 발전량 (시뮬레이션, MWh)
PV0_MWH = {
    1:      0.0,
    "2_70": 21.481, "2_80": 20.054, "2_90": 18.177,
    3:     35.471,
}

CASES = [1, "2_70", "2_80", "2_90", 3]

# Case 2 계열은 동일한 경제 파라미터 사용
capex_map    = {1: capex[1], "2_70": capex[2], "2_80": capex[2], "2_90": capex[2], 3: capex[3]}
om_yr1_map   = {1: om_yr1[1], "2_70": om_yr1[2], "2_80": om_yr1[2], "2_90": om_yr1[2], 3: om_yr1[3]}
rep_year_map = {1: rep_year[1], "2_70": rep_year[2], "2_80": rep_year[2], "2_90": rep_year[2], 3: rep_year[3]}
rep_cost_map = {1: rep_cost[1], "2_70": rep_cost[2], "2_80": rep_cost[2], "2_90": rep_cost[2], 3: rep_cost[3]}

# ── 연도별 비용 계산 (공통) ───────────────────────────────────────────
years    = np.arange(0, n_years + 1)
r_disc   = eval_cell(ws3, 12)   # 할인율 0.055

annual_cost = {k: np.zeros(n_years + 1) for k in CASES}  # 연도별 명목 비용
for k in CASES:
    annual_cost[k][0] = capex_map[k]

for t in range(1, n_years + 1):
    tariff_t = tariff_0 * (1 + r_elec) ** (t - 1)
    for k in CASES:
        pv_t      = PV0_MWH[k] * (1 - r_deg) ** (t - 1)
        net_pv    = pv_t - (drive_3_kwh / 1000.0 if k == 3 else 0.0)
        net_kWh   = (HVAC_MWH[k] - net_pv) * 1000.0
        elec_cost = net_kWh * tariff_t
        om_t      = om_yr1_map[k] * (1 + r_inf) ** (t - 1)
        rep_t     = rep_cost_map[k] * (1 + r_inf) ** t if t == rep_year_map[k] else 0.0
        annual_cost[k][t] = elec_cost + om_t + rep_t

# ── 명목 누적 비용 (Nominal Cumulative Cost) ─────────────────────────
cum_nominal = {k: np.zeros(n_years + 1) for k in CASES}
for k in CASES:
    cum_nominal[k][0] = annual_cost[k][0]
    for t in range(1, n_years + 1):
        cum_nominal[k][t] = cum_nominal[k][t - 1] + annual_cost[k][t]

# ── LCC: 할인된 누적 비용 (Present Value) ────────────────────────────
cum_lcc = {k: np.zeros(n_years + 1) for k in CASES}
for k in CASES:
    cum_lcc[k][0] = annual_cost[k][0]   # CAPEX는 year 0, 할인 없음
    for t in range(1, n_years + 1):
        cum_lcc[k][t] = cum_lcc[k][t - 1] + annual_cost[k][t] / (1 + r_disc) ** t

# 백만원 단위
for k in CASES:
    cum_nominal[k] /= 1e6
    cum_lcc[k]     /= 1e6

print(f"\n25년 명목 누적 비용:")
for k in CASES:
    print(f"  {k}: {cum_nominal[k][25]:.1f} M KRW")
print(f"\n25년 LCC (현재가치):")
for k in CASES:
    print(f"  {k}: {cum_lcc[k][25]:.1f} M KRW")

# ── 공통 플롯 스타일 ──────────────────────────────────────────────────
dm.style.use("presentation")
plt.rcParams.update({"xtick.labelsize": 12, "ytick.labelsize": 12})

colors  = {1: "oc.gray7", "2_70": "oc.blue4", "2_80": "oc.blue5", "2_90": "oc.blue7", 3: "oc.orange7"}
markers = {1: "o",        "2_70": "D",         "2_80": "s",        "2_90": "o",        3: "^"}
labels  = {
    1:      "Case 1 (Base, No BIPV)",
    "2_70": "Case 2 — Fixed 70°",
    "2_80": "Case 2 — Fixed 80°",
    "2_90": "Case 2 — Fixed 90°",
    3:      "Case 3 (Kinetic BIPV)",
}
lws = {1: 1.5, "2_70": 1.3, "2_80": 1.3, "2_90": 1.3, 3: 1.5}


def compute_label_positions(cum, min_gap=8.0):
    """
    양방향 relaxation: 위아래로 밀어내며 수렴.
    정렬 순서는 고정하고 인접 쌍이 min_gap 미만이면 절반씩 밀어냄.
    """
    sorted_cases = sorted(CASES, key=lambda k: cum[k][25])
    pos = {k: float(cum[k][25]) for k in sorted_cases}

    for _ in range(300):
        changed = False
        for i in range(1, len(sorted_cases)):
            k, k_prev = sorted_cases[i], sorted_cases[i - 1]
            gap = pos[k] - pos[k_prev]
            if gap < min_gap:
                push = (min_gap - gap) / 2
                pos[k_prev] -= push
                pos[k]      += push
                changed = True
        if not changed:
            break
    return pos


def draw_cost_plot(ax, cum, ylabel, title):
    for k in CASES:
        ax.plot(years, cum[k], color=colors[k], lw=dm.lw(lws[k]),
                marker=markers[k], markersize=3, label=labels[k])

    # 레이블 위치 자동 계산 → ylim을 레이블 최고점 기준으로 설정
    label_y = compute_label_positions(cum)
    y_top   = max(max(cum[k].max() for k in CASES), max(label_y.values())) * 1.06
    ax.set_ylim(0, y_top)

    for k in CASES:
        ax.annotate(f"{cum[k][25]:.1f} M",
                    xy=(25, cum[k][25]), xytext=(25.5, label_y[k]),
                    fontsize=9.5, color=colors[k], fontweight="bold", va="center",
                    arrowprops=dict(arrowstyle="-", color=colors[k], lw=0.9))

    # 12년차 교체비 이벤트
    ax.axvline(12, color="oc.gray5", linestyle=":", lw=dm.lw(0.8))
    ax.text(12.3, y_top * 0.03,
            "Year 12\n(Actuator\nReplacement)",
            color="oc.gray6", fontsize=9, va="bottom")

    ax.set_xlabel("Project Timeline [Years]")
    ax.set_ylabel(ylabel)
    ax.set_title(title, fontweight="bold")
    ax.set_xlim(0, 28)
    ax.set_xticks(range(0, 26, 5))
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}"))
    ax.legend(loc="upper left")


# ── 그래프 1: 명목 누적 비용 ─────────────────────────────────────────
fig, ax = plt.subplots(figsize=(23 / 2.54, 14 / 2.54))
draw_cost_plot(
    ax, cum_nominal,
    ylabel="Nominal Cumulative Cost [Million KRW]",
    title="Nominal Cumulative Cost Comparison\n"
          "(CAPEX + HVAC Electricity + O&M + Replacement, Undiscounted)",
)
dm.simple_layout(fig)
fig.savefig(os.path.join(figure_dir, "fig8_cumulative_nominal.png"), dpi=300, transparent=True)
fig.savefig(os.path.join(figure_dir, "fig8_cumulative_nominal.svg"), transparent=True)
plt.close(fig)

# ── 그래프 2: LCC (현재가치 기준) ────────────────────────────────────
fig, ax = plt.subplots(figsize=(23 / 2.54, 14 / 2.54))
draw_cost_plot(
    ax, cum_lcc,
    ylabel="Life Cycle Cost — Present Value [Million KRW]",
    title=f"Life Cycle Cost (LCC) Comparison\n"
          f"(Discount Rate {r_disc*100:.1f}%,  Tariff Escalation {r_elec*100:.0f}%/yr,  PV Degradation {r_deg*100:.1f}%/yr)",
)

# 저장 처리
output_png = os.path.join(figure_dir, "fig9_cumulative_lcc.png")
output_svg = os.path.join(figure_dir, "fig9_cumulative_lcc.svg")

dm.simple_layout(fig)
fig.savefig(output_png, dpi=300, transparent=True)
fig.savefig(output_svg, transparent=True)
plt.close(fig)

print(f"\n저장 완료: {figure_dir}")
