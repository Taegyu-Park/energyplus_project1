"""
DCF-based Investment Recovery Comparison: Case 2 vs Case 3
현금흐름할인법(DCF) 투자비용 회수 비교 — 정적 BIPV 90° vs Kinetic BIPV

Parameters from economy_analysis.xlsx (가정 sheets)
PV generation from EnergyPlus simulation
"""

import os
import re
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import dartwork_mpl as dm
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ── 경로 ─────────────────────────────────────────────────────────────
script_dir   = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.normpath(os.path.join(script_dir, "..", "..", ".."))
figure_dir   = os.path.normpath(os.path.join(script_dir, "..", "..", "figure", "LLC_dcf"))
excel_path   = os.path.join(project_root, "economy_analysis.xlsx")
os.makedirs(figure_dir, exist_ok=True)

# ── Excel 파라미터 읽기 ───────────────────────────────────────────────
wb  = load_workbook(excel_path, data_only=False)
ws3 = wb["가정"]
ws2 = wb["가정_Case2"]

def ec(ws, row, col=2):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str) and v.startswith("="):
        expr = v[1:]
        m = re.fullmatch(r'([A-Z]+)(\d+)', expr)
        if m:
            return ec(ws, int(m.group(2)), column_index_from_string(m.group(1)))
        try: return float(eval(expr))
        except: return 0.0
    return 0.0 if v is None else float(v)

# 공통 경제 파라미터
b42, b43, b44, b45 = ec(ws3,42), ec(ws3,43), ec(ws3,44), ec(ws3,45)
d34, d35, d36       = ec(ws3,34,4), ec(ws3,35,4), ec(ws3,36,4)
f = 1 + b44 + b45
tariff_0 = ((132.4+b42+b43)*f*d34 + (91.9+b42+b43)*f*d35 + (119+b42+b43)*f*d36)

r_elec  = ec(ws3, 10)   # 0.05
r_inf   = ec(ws3, 11)   # 0.02
r_deg   = ec(ws3, 7)    # 0.005
r_disc  = ec(ws3, 12)   # 0.055
n_years = int(ec(ws3, 8))  # 25

# CAPEX & 운영비
capex2  = sum(ec(ws2, r) for r in range(15, 23))
capex3  = sum(ec(ws3, r) for r in range(15, 23))
om_r2   = ec(ws2, 26)
om_r3   = ec(ws3, 26)
rep_yr3 = int(ec(ws3, 28))
rep_c3  = ec(ws3, 29)
drive3  = ec(ws3, 6)   # kWh/yr

# EnergyPlus 시뮬레이션 발전량 (1년차) — 전 각도
ANGLES = list(range(0, 100, 10))
PV2_kwh = {
     0: 22_927.0, 10: 23_341.0, 20: 22_985.0, 30: 23_476.0,
    40: 23_413.0, 50: 23_012.0, 60: 22_445.0, 70: 21_481.0,
    80: 20_054.0, 90: 18_177.0,
}
PV3_kwh = 35_471.0

# ── DCF 계산 ─────────────────────────────────────────────────────────
def calc_dcf(capex, pv0_kwh, drive_kwh, om_rate, rep_year, rep_cost_now):
    cum  = np.zeros(n_years + 1)
    cum[0] = -capex
    om0  = capex * om_rate
    for t in range(1, n_years + 1):
        pv_t    = pv0_kwh * (1 - r_deg) ** (t - 1)
        net_pv  = pv_t - drive_kwh
        savings = net_pv  * tariff_0 * (1 + r_elec) ** (t - 1)
        om_t    = om0     * (1 + r_inf)  ** (t - 1)
        rep_t   = rep_cost_now * (1 + r_inf) ** t if t == rep_year else 0.0
        cum[t]  = cum[t - 1] + (savings - om_t - rep_t) / (1 + r_disc) ** t
    return cum

dcf2   = {a: calc_dcf(capex2, PV2_kwh[a], 0.0, om_r2, 99, 0.0) for a in ANGLES}
dcf3   = calc_dcf(capex3, PV3_kwh, drive3, om_r3, rep_yr3, rep_c3)

years  = np.arange(n_years + 1)
dcf2_M = {a: dcf2[a] / 1e6 for a in ANGLES}
dcf3_M = dcf3 / 1e6

def payback_yr(arr):
    for i in range(1, len(arr)):
        if arr[i - 1] < 0 <= arr[i]:
            return (i - 1) + abs(arr[i - 1]) / (arr[i] - arr[i - 1])
    return None

pb2  = {a: payback_yr(dcf2_M[a]) for a in ANGLES}
pb3  = payback_yr(dcf3_M)
best_a = max(ANGLES, key=lambda a: dcf2_M[a][-1])   # 최고 NPV 각도

for a in ANGLES:
    pb = pb2[a]
    print(f"[Case 2 {a:2d}°] PV={PV2_kwh[a]:.0f}kWh  "
          f"Payback={'> 25 yr' if pb is None else f'{pb:.1f} yr':>8s}  "
          f"25yr NPV={dcf2_M[a][-1]:+.2f} M KRW")
print(f"[Case 3    ] PV={PV3_kwh:.0f}kWh  "
      f"Payback={'> 25 yr' if pb3 is None else f'{pb3:.1f} yr':>8s}  "
      f"25yr NPV={dcf3_M[-1]:+.2f} M KRW")

# ── 플롯 ─────────────────────────────────────────────────────────────
import matplotlib.cm as cm
dm.style.use("presentation")
plt.rcParams.update({"xtick.labelsize": 12, "ytick.labelsize": 12})

fig, ax = plt.subplots(figsize=dm.figsize("26cm", "14cm"))

cmap      = cm.Blues
C2_color  = {a: cmap(0.25 + 0.65 * i / (len(ANGLES) - 1)) for i, a in enumerate(ANGLES)}
C3        = "oc.orange7"

# ── y축 범위 먼저 계산 ──────────────────────────────────────────────
all_vals = np.concatenate(list(dcf2_M.values()) + [dcf3_M])
y_lo = all_vals.min() * 1.12
y_hi = all_vals.max() * 1.22
ax.set_ylim(y_lo, y_hi)

# ── 음수/양수 음영 ─────────────────────────────────────────────────
ax.fill_between(years, y_lo, 0, color="oc.red2",  alpha=0.08, linewidth=0, zorder=0)
ax.fill_between(years, 0, y_hi, color="oc.teal1", alpha=0.10, linewidth=0, zorder=0)
ax.axhline(0, color="oc.gray5", linestyle="--", lw=dm.lw(0.8), zorder=1)

# ── Case 2 곡선 (0°–90°, 그라디언트) ──────────────────────────────
for a in ANGLES:
    ax.plot(years, dcf2_M[a], color=C2_color[a], lw=dm.lw(1.3),
            label=f"Case 2 — Fixed {a}°", zorder=3)

# ── Case 3 곡선 ─────────────────────────────────────────────────────
ax.plot(years, dcf3_M, color=C3, lw=dm.lw(2.2),
        label=f"Case 3 — Kinetic BIPV", zorder=4)

# ── Case 3 회수기간 ─────────────────────────────────────────────────
if pb3 is not None:
    ax.axvline(pb3, color=C3, linestyle=":", lw=dm.lw(1.2), zorder=2)
    ax.plot(pb3, 0, marker="o", color=C3, markersize=7, zorder=5)
    ax.text(pb3 - 0.3, y_hi * 0.88,
            f"Case 3 Payback\n{pb3:.1f} yr",
            color=C3, fontsize=10.5, fontweight="bold", ha="right", va="top")

# ── Case 3 12년차 교체비 ────────────────────────────────────────────
yr12_val = dcf3_M[12]
ax.plot(12, yr12_val, marker="x", color=C3, markersize=9, mew=2, zorder=5)
ax.annotate("Actuator\nReplacement\n(Year 12)",
            xy=(12, yr12_val), xytext=(8.5, -8),
            fontsize=9.5, color="oc.orange9", ha="right", va="center",
            arrowprops=dict(arrowstyle="->", color="oc.orange9", lw=1.1))

# ── 우측 끝값 레이블: Case 3 + 최고/최저 Case 2만 표시 ──────────────
worst_a = 90
ax.text(25.3, dcf3_M[-1] + 0.5,
        f"+{dcf3_M[-1]:.1f} M  (Case 3)",
        color=C3, fontsize=10.5, fontweight="bold", va="center")

# best 각도 (30°)
pb_best = pb2[best_a]
suffix_best = f"  ({pb_best:.1f} yr)" if pb_best else ""
ax.annotate(f"{dcf2_M[best_a][-1]:+.1f} M  ({best_a}°){suffix_best}",
            xy=(25, dcf2_M[best_a][-1]),
            xytext=(25.3, dcf2_M[best_a][-1] + 2.5),
            fontsize=9.5, color=C2_color[best_a], fontweight="bold", va="center",
            arrowprops=dict(arrowstyle="-", color=C2_color[best_a], lw=0.8))

# worst 각도 (90°)
pb_worst = pb2[worst_a]
suffix_worst = "  (> 25 yr)" if pb_worst is None else f"  ({pb_worst:.1f} yr)"
ax.annotate(f"{dcf2_M[worst_a][-1]:+.1f} M  ({worst_a}°){suffix_worst}",
            xy=(25, dcf2_M[worst_a][-1]),
            xytext=(25.3, dcf2_M[worst_a][-1] - 2.5),
            fontsize=9.5, color=C2_color[worst_a], fontweight="bold", va="center",
            arrowprops=dict(arrowstyle="-", color=C2_color[worst_a], lw=0.8))

# ── 축/제목 ────────────────────────────────────────────────────────
ax.set_xlabel("Project Timeline [Years]")
ax.set_ylabel("Cumulative NPV [Million KRW]")
ax.set_title(
    "DCF-based Investment Recovery Comparison\n"
    f"(Discount Rate {r_disc*100:.1f}%,  Tariff Escalation {r_elec*100:.0f}%/yr,  PV Degradation {r_deg*100:.1f}%/yr)",
    fontweight="bold",
)
ax.set_xlim(0, 31)
ax.set_xticks(range(0, 26, 5))
ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}"))
ax.legend(loc="lower right", fontsize=9, ncol=2)

dm.simple_layout(fig)
fig.savefig(os.path.join(figure_dir, "dcf_payback_comparison.png"), dpi=300, transparent=True)
fig.savefig(os.path.join(figure_dir, "dcf_payback_comparison.svg"), transparent=True)
plt.close(fig)
print(f"\nSaved → {figure_dir}/dcf_payback_comparison.{{png,svg}}")
