"""
Cumulative Cost Comparison over 25-year project life
  - cumulative_nominal.svg : undiscounted (nominal) sum
  - cumulative_lcc.svg     : discounted to present value (LCC)
"""

import os
import re
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.patches as mpatches
import dartwork_mpl as dm
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ── 경로 ─────────────────────────────────────────────────────────────
script_dir   = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.normpath(os.path.join(script_dir, "..", "..", ".."))
figure_dir   = os.path.normpath(os.path.join(script_dir, "..", "..", "figure", "cumulative_cost"))
excel_path   = os.path.join(project_root, "economy_analysis.xlsx")
os.makedirs(figure_dir, exist_ok=True)

# ── Excel 파라미터 ────────────────────────────────────────────────────
wb  = load_workbook(excel_path, data_only=False)
ws3 = wb["가정"]
ws2 = wb["가정_Case2"]

def ec(ws, row, col=2):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, (int, float)): return float(v)
    if isinstance(v, str) and v.startswith("="):
        expr = v[1:]
        m = re.fullmatch(r'([A-Z]+)(\d+)', expr)
        if m:
            return ec(ws, int(m.group(2)), column_index_from_string(m.group(1)))
        try: return float(eval(expr))
        except: return 0.0
    return 0.0 if v is None else float(v)

b42, b43, b44, b45 = ec(ws3,42), ec(ws3,43), ec(ws3,44), ec(ws3,45)
d34, d35, d36       = ec(ws3,34,4), ec(ws3,35,4), ec(ws3,36,4)
f = 1 + b44 + b45
tariff_0 = ((132.4+b42+b43)*f*d34 + (91.9+b42+b43)*f*d35 + (119+b42+b43)*f*d36)

r_elec  = ec(ws3, 10)
r_inf   = ec(ws3, 11)
r_deg   = ec(ws3, 7)
r_disc  = ec(ws3, 12)
n_years = int(ec(ws3, 8))

capex2  = sum(ec(ws2, r) for r in range(15, 23))
capex3  = sum(ec(ws3, r) for r in range(15, 23))
om_r2   = ec(ws2, 26)
om_r3   = ec(ws3, 26)
rep_yr3 = int(ec(ws3, 28))
rep_c3  = ec(ws3, 29)
drive3  = ec(ws3, 6)

print(f"tariff_0      = {tariff_0:.2f} won/kWh")
print(f"Case 2 CAPEX  = {capex2/1e6:.1f} M,  O&M = {om_r2*100:.1f}%/yr")
print(f"Case 3 CAPEX  = {capex3/1e6:.1f} M,  O&M = {om_r3*100:.1f}%/yr")
print(f"Case 3 rep    = year {rep_yr3}, {rep_c3/1e6:.0f} M (present)")

# ── 에너지 데이터 (EnergyPlus 시뮬레이션, MWh/yr) ────────────────────
CASES    = [1, "2_70", "2_80", "2_90", 3]
HVAC_MWH = {1: 48.994, "2_70": 51.460, "2_80": 50.994, "2_90": 50.110, 3: 48.722}
PV0_MWH  = {1: 0.0,    "2_70": 21.481, "2_80": 20.054, "2_90": 18.177, 3: 35.471}

def capex_of(k):
    if k == 1:       return 0.0
    if k == 3:       return capex3
    return capex2

def om_rate_of(k):
    if k == 1:       return 0.0
    if k == 3:       return om_r3
    return om_r2

# ── 연간 비용 계산 ────────────────────────────────────────────────────
years = np.arange(n_years + 1)

annual_cost = {k: np.zeros(n_years + 1) for k in CASES}
for k in CASES:
    cap   = capex_of(k)
    omr   = om_rate_of(k)
    om0   = cap * omr
    annual_cost[k][0] = cap   # CAPEX at t=0

    for t in range(1, n_years + 1):
        pv_t       = PV0_MWH[k] * 1000 * (1 - r_deg) ** (t - 1)   # kWh
        drv_t      = drive3 if k == 3 else 0.0                      # kWh
        net_kwh    = HVAC_MWH[k] * 1000 - pv_t + drv_t
        elec_t     = net_kwh * tariff_0 * (1 + r_elec) ** (t - 1)
        om_t       = om0 * (1 + r_inf) ** (t - 1)
        rep_t      = rep_c3 * (1 + r_inf) ** t if (k == 3 and t == rep_yr3) else 0.0
        annual_cost[k][t] = elec_t + om_t + rep_t

# ── 누적 비용 계산 ────────────────────────────────────────────────────
# Nominal: 할인 없이 그냥 합산
cum_nom = {k: np.zeros(n_years + 1) for k in CASES}
for k in CASES:
    cum_nom[k][0] = annual_cost[k][0]
    for t in range(1, n_years + 1):
        cum_nom[k][t] = cum_nom[k][t-1] + annual_cost[k][t]

# LCC: 미래 비용을 현재가치로 할인
cum_lcc = {k: np.zeros(n_years + 1) for k in CASES}
for k in CASES:
    cum_lcc[k][0] = annual_cost[k][0]   # CAPEX: 현재이므로 할인 없음
    for t in range(1, n_years + 1):
        cum_lcc[k][t] = cum_lcc[k][t-1] + annual_cost[k][t] / (1 + r_disc) ** t

print("\n25년 누적 명목 비용:")
for k in CASES:
    print(f"  {k}: {cum_nom[k][25]/1e6:.1f} M KRW")
print("\n25년 LCC (현재가치):")
for k in CASES:
    print(f"  {k}: {cum_lcc[k][25]/1e6:.1f} M KRW")

# ── 스타일 ────────────────────────────────────────────────────────────
dm.style.use("presentation")
plt.rcParams.update({"xtick.labelsize": 11, "ytick.labelsize": 11})

colors  = {1: "oc.gray6", "2_70": "oc.blue4", "2_80": "oc.blue6", "2_90": "oc.blue8", 3: "oc.orange6"}
lws     = {1: 1.8, "2_70": 1.2, "2_80": 1.2, "2_90": 1.2, 3: 2.2}
markers = {1: "o", "2_70": None, "2_80": None, "2_90": None, 3: "s"}
labels  = {1: "Case 1 (Base)",
           "2_70": "Case 2 — 70°",
           "2_80": "Case 2 — 80°",
           "2_90": "Case 2 — 90°",
           3: "Case 3 (Kinetic)"}

# ── 레이블 위치: 양방향 relaxation ────────────────────────────────────
def compute_label_positions(cum, min_gap=8.0):
    sorted_cases = sorted(CASES, key=lambda k: cum[k][25])
    pos = {k: float(cum[k][25]) / 1e6 for k in sorted_cases}

    for _ in range(300):
        changed = False
        for i in range(1, len(sorted_cases)):
            k, k_prev = sorted_cases[i], sorted_cases[i - 1]
            gap = pos[k] - pos[k_prev]
            if gap < min_gap:
                push = (min_gap - gap) / 2
                pos[k_prev] -= push
                pos[k]      += push
                changed = True
        if not changed:
            break
    return pos

def draw_cost_plot(ax, cum, ylabel, title, min_gap=8.0):
    cum_M = {k: cum[k] / 1e6 for k in CASES}
    for k in CASES:
        mk = dict(marker=markers[k], markersize=3) if markers[k] else {}
        ax.plot(years, cum_M[k], color=colors[k], lw=dm.lw(lws[k]),
                label=labels[k], **mk)

    label_y = compute_label_positions(cum, min_gap)
    y_top   = max(max(cum_M[k].max() for k in CASES), max(label_y.values())) * 1.06
    ax.set_ylim(0, y_top)

    for k in CASES:
        ax.annotate(f"{cum_M[k][25]:.1f} M",
                    xy=(25, cum_M[k][25]), xytext=(25.5, label_y[k]),
                    fontsize=9.5, color=colors[k], fontweight="bold", va="center",
                    arrowprops=dict(arrowstyle="-", color=colors[k], lw=0.9))

    ax.axvline(12, color="oc.gray5", linestyle=":", lw=dm.lw(0.8))
    ax.text(12.3, y_top * 0.03, "Year 12\n(Actuator\nReplacement)",
            color="oc.gray6", fontsize=9, va="bottom")
    ax.set_xlabel("Project Timeline [Years]")
    ax.set_ylabel(ylabel)
    ax.set_title(title, fontweight="bold")
    ax.set_xlim(0, 28)
    ax.set_xticks(range(0, 26, 5))
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{x:.0f}"))
    ax.legend(loc="upper left")

# ── 그래프 저장 ───────────────────────────────────────────────────────
for cum, fname, ylabel, title in [
    (cum_nom,
     "cumulative_nominal",
     "Nominal Cumulative Cost [Million KRW]",
     "25-Year Nominal Cumulative Cost Comparison\n(Undiscounted)"),
    (cum_lcc,
     "cumulative_lcc",
     "Life Cycle Cost [Million KRW]",
     "25-Year Life Cycle Cost (LCC) Comparison\n(Discounted at 5.5%)"),
]:
    fig, ax = plt.subplots(figsize=dm.figsize("23cm", "14cm"))
    draw_cost_plot(ax, cum, ylabel, title)
    dm.simple_layout(fig)
    fig.savefig(os.path.join(figure_dir, f"{fname}.png"), dpi=300, bbox_inches="tight", transparent=True)
    fig.savefig(os.path.join(figure_dir, f"{fname}.svg"), bbox_inches="tight", transparent=True)
    plt.close(fig)

print(f"\n저장 완료: {figure_dir}")
